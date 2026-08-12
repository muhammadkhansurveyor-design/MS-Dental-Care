
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
from datetime import datetime, date
from io import BytesIO
import os, csv, shutil, zipfile
from openpyxl import load_workbook

app = Flask(__name__)
app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "CHANGE_ME_IN_PRODUCTION")
db_url = os.environ.get("DATABASE_URL", "sqlite:///ms_dental_final.db")
if db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config["UPLOAD_FOLDER"] = os.path.join(os.path.dirname(__file__), "static", "uploads")
app.config["MAX_CONTENT_LENGTH"] = 8 * 1024 * 1024
db = SQLAlchemy(app)

ROLE_SUPER = "super_admin"
ROLE_BRANCH = "branch_admin"
ROLE_ENTRY = "entry_user"
ROLE_VIEWER = "viewer"

class Branch(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(120),nullable=False)
    country=db.Column(db.String(80),default="Pakistan")
    city=db.Column(db.String(80),default="")
    address=db.Column(db.String(255),default="")
    phone=db.Column(db.String(50),default="")
    active=db.Column(db.Boolean,default=True)

class User(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(120),nullable=False)
    email=db.Column(db.String(180),unique=True,nullable=False)
    password_hash=db.Column(db.String(255),nullable=False)
    role=db.Column(db.String(30),nullable=False)
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=True)
    active=db.Column(db.Boolean,default=True)
    branch=db.relationship("Branch")

class Doctor(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(120),nullable=False)
    mobile=db.Column(db.String(40),default="")
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=True)
    active=db.Column(db.Boolean,default=True)
    branch=db.relationship("Branch")

class Staff(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(120),nullable=False)
    role_name=db.Column(db.String(80),default="")
    mobile=db.Column(db.String(40),default="")
    cnic=db.Column(db.String(40),default="")
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=False)
    salary=db.Column(db.Float,default=0)
    active=db.Column(db.Boolean,default=True)
    branch=db.relationship("Branch")

class SalaryPayment(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    staff_id=db.Column(db.Integer,db.ForeignKey("staff.id"),nullable=False)
    month=db.Column(db.String(20),nullable=False)
    gross_salary=db.Column(db.Float,default=0)
    advance=db.Column(db.Float,default=0)
    paid=db.Column(db.Float,default=0)
    balance=db.Column(db.Float,default=0)
    paid_date=db.Column(db.Date,nullable=True)
    staff=db.relationship("Staff")

class Treatment(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    name=db.Column(db.String(160),unique=True,nullable=False)
    price=db.Column(db.Float,default=0)
    active=db.Column(db.Boolean,default=True)

class Patient(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    patient_code=db.Column(db.String(30),unique=True)
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=False)
    name=db.Column(db.String(160),nullable=False)
    mobile=db.Column(db.String(40),default="")
    age=db.Column(db.String(20),default="")
    gender=db.Column(db.String(20),default="")
    address=db.Column(db.String(255),default="")
    medical_history=db.Column(db.Text,default="")
    created_by=db.Column(db.Integer,db.ForeignKey("user.id"))
    created_at=db.Column(db.DateTime,default=datetime.utcnow)
    branch=db.relationship("Branch")

class Visit(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    patient_id=db.Column(db.Integer,db.ForeignKey("patient.id"),nullable=False)
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=False)
    doctor_id=db.Column(db.Integer,db.ForeignKey("doctor.id"),nullable=True)
    treatment_id=db.Column(db.Integer,db.ForeignKey("treatment.id"),nullable=True)
    tooth_notes=db.Column(db.String(255),default="")
    clinical_notes=db.Column(db.Text,default="")
    xray_file=db.Column(db.String(255),default="")
    visit_date=db.Column(db.Date,default=date.today)
    created_by=db.Column(db.Integer,db.ForeignKey("user.id"))
    patient=db.relationship("Patient")
    doctor=db.relationship("Doctor")
    treatment=db.relationship("Treatment")
    branch=db.relationship("Branch")

class Invoice(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    invoice_no=db.Column(db.String(30),unique=True)
    patient_id=db.Column(db.Integer,db.ForeignKey("patient.id"),nullable=False)
    visit_id=db.Column(db.Integer,db.ForeignKey("visit.id"),nullable=True)
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=False)
    total=db.Column(db.Float,default=0)
    discount=db.Column(db.Float,default=0)
    received=db.Column(db.Float,default=0)
    balance=db.Column(db.Float,default=0)
    payment_type=db.Column(db.String(30),default="Cash")
    created_by=db.Column(db.Integer,db.ForeignKey("user.id"))
    created_at=db.Column(db.DateTime,default=datetime.utcnow)
    patient=db.relationship("Patient")
    visit=db.relationship("Visit")
    branch=db.relationship("Branch")

class Payment(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    invoice_id=db.Column(db.Integer,db.ForeignKey("invoice.id"),nullable=False)
    amount=db.Column(db.Float,default=0)
    payment_date=db.Column(db.Date,default=date.today)
    received_by=db.Column(db.Integer,db.ForeignKey("user.id"))
    invoice=db.relationship("Invoice")

class Expense(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=False)
    expense_date=db.Column(db.Date,default=date.today)
    category=db.Column(db.String(120),default="")
    description=db.Column(db.String(255),default="")
    amount=db.Column(db.Float,default=0)
    created_by=db.Column(db.Integer,db.ForeignKey("user.id"))
    branch=db.relationship("Branch")

class ActivityLog(db.Model):
    id=db.Column(db.Integer,primary_key=True)
    user_id=db.Column(db.Integer,db.ForeignKey("user.id"),nullable=True)
    branch_id=db.Column(db.Integer,db.ForeignKey("branch.id"),nullable=True)
    action=db.Column(db.String(120),nullable=False)
    details=db.Column(db.Text,default="")
    created_at=db.Column(db.DateTime,default=datetime.utcnow)
    user=db.relationship("User")
    branch=db.relationship("Branch")

def current_user():
    uid=session.get("user_id")
    return db.session.get(User,uid) if uid else None

def login_required(fn):
    @wraps(fn)
    def w(*a,**k):
        if not current_user(): return redirect(url_for("login"))
        return fn(*a,**k)
    return w

def roles_required(*roles):
    def deco(fn):
        @wraps(fn)
        def w(*a,**k):
            u=current_user()
            if not u or u.role not in roles:
                flash("Permission denied.","danger"); return redirect(url_for("dashboard"))
            return fn(*a,**k)
        return w
    return deco

def super_required(fn): return roles_required(ROLE_SUPER)(fn)

def allowed_branch_ids(u):
    if u.role==ROLE_SUPER: return [b.id for b in Branch.query.filter_by(active=True).all()]
    return [u.branch_id] if u.branch_id else []

def assert_branch(branch_id):
    return branch_id in allowed_branch_ids(current_user())

def log_action(action,details="",branch_id=None):
    u=current_user()
    db.session.add(ActivityLog(user_id=u.id if u else None,branch_id=branch_id,action=action,details=details))
    db.session.commit()

@app.context_processor
def inject(): return dict(current_user=current_user())


EXCEL_TREATMENT_COLUMNS = {
    8: "Consultation",
    9: "Composite Filling",
    10: "GIC Filling",
    11: "Scaling & Polishing",
    12: "Tooth Extraction",
    13: "Surgical Extraction",
    14: "Root Canal Treatment",
    15: "Crown Zirconia",
    16: "Crown",
    17: "Bridge",
    18: "Denture",
    19: "Implant",
    20: "Braces (Orthodontic Treatment)",
    21: "X-Ray",
    22: "Other",
}

def excel_date(value):
    if hasattr(value, "date"):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str) and value.strip():
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    return date.today()

def get_or_create_branch(name):
    name = (name or "Main Branch").strip()
    b = Branch.query.filter(db.func.lower(Branch.name) == name.lower()).first()
    if not b:
        b = Branch(name=name, country="Pakistan")
        db.session.add(b)
        db.session.flush()
    return b

def get_or_create_treatment(name):
    t = Treatment.query.filter(db.func.lower(Treatment.name) == name.lower()).first()
    if not t:
        t = Treatment(name=name, price=0)
        db.session.add(t)
        db.session.flush()
    return t


@app.route("/",methods=["GET","POST"])
def login():
    if request.method=="POST":
        u=User.query.filter_by(email=request.form["email"].strip().lower(),active=True).first()
        if u and check_password_hash(u.password_hash,request.form["password"]):
            session["user_id"]=u.id; log_action("LOGIN",u.email,u.branch_id); return redirect(url_for("dashboard"))
        flash("Invalid email or password.","danger")
    return render_template("login.html")

@app.route("/logout")
@login_required
def logout():
    u=current_user(); log_action("LOGOUT",u.email,u.branch_id); session.clear(); return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    u=current_user(); bids=allowed_branch_ids(u)
    patients=Patient.query.filter(Patient.branch_id.in_(bids)).count() if bids else 0
    invs=Invoice.query.filter(Invoice.branch_id.in_(bids)).all() if bids else []
    exps=Expense.query.filter(Expense.branch_id.in_(bids)).all() if bids else []
    billing=sum(i.total-i.discount for i in invs); received=sum(i.received for i in invs); due=sum(i.balance for i in invs); expenses=sum(e.amount for e in exps)
    rows=[]
    for b in Branch.query.filter(Branch.id.in_(bids)).all() if bids else []:
        bi=Invoice.query.filter_by(branch_id=b.id).all(); be=Expense.query.filter_by(branch_id=b.id).all()
        rows.append(dict(name=b.name,patients=Patient.query.filter_by(branch_id=b.id).count(),
                         billing=sum(i.total-i.discount for i in bi),received=sum(i.received for i in bi),
                         balance=sum(i.balance for i in bi),expenses=sum(e.amount for e in be),
                         net=sum(i.received for i in bi)-sum(e.amount for e in be)))
    return render_template("dashboard.html",patients=patients,billing=billing,received=received,due=due,expenses=expenses,net=received-expenses,rows=rows)

@app.route("/patients")
@login_required
def patients():
    u=current_user(); bids=allowed_branch_ids(u); q=request.args.get("q","").strip()
    query=Patient.query.filter(Patient.branch_id.in_(bids)) if bids else Patient.query.filter(False)
    if q: query=query.filter((Patient.name.ilike(f"%{q}%"))|(Patient.mobile.ilike(f"%{q}%"))|(Patient.patient_code.ilike(f"%{q}%")))
    return render_template("patients.html",rows=query.order_by(Patient.id.desc()).limit(500).all(),q=q)

@app.route("/patients/new",methods=["GET","POST"])
@login_required
@roles_required(ROLE_SUPER,ROLE_BRANCH,ROLE_ENTRY)
def patient_new():
    u=current_user(); bids=allowed_branch_ids(u)
    branches=Branch.query.filter(Branch.id.in_(bids)).all(); doctors=Doctor.query.filter(Doctor.branch_id.in_(bids),Doctor.active==True).all(); treatments=Treatment.query.filter_by(active=True).order_by(Treatment.name).all()
    if request.method=="POST":
        branch_id=int(request.form["branch_id"])
        if not assert_branch(branch_id): flash("Branch access denied.","danger"); return redirect(url_for("patient_new"))
        p=Patient(branch_id=branch_id,name=request.form["name"].strip(),mobile=request.form.get("mobile",""),age=request.form.get("age",""),gender=request.form.get("gender",""),
                  address=request.form.get("address",""),medical_history=request.form.get("medical_history",""),created_by=u.id)
        db.session.add(p); db.session.flush(); p.patient_code=f"MS-P-{p.id:05d}"
        xray=""
        f=request.files.get("xray")
        if f and f.filename:
            fn=f"{p.id}_{int(datetime.utcnow().timestamp())}_{secure_filename(f.filename)}"; f.save(os.path.join(app.config["UPLOAD_FOLDER"],fn)); xray=fn
        visit=Visit(patient_id=p.id,branch_id=branch_id,doctor_id=int(request.form["doctor_id"]) if request.form.get("doctor_id") else None,
                    treatment_id=int(request.form["treatment_id"]) if request.form.get("treatment_id") else None,tooth_notes=request.form.get("tooth_notes",""),
                    clinical_notes=request.form.get("clinical_notes",""),xray_file=xray,created_by=u.id)
        db.session.add(visit); db.session.flush()
        total=float(request.form.get("total") or 0); discount=float(request.form.get("discount") or 0); rec=float(request.form.get("received") or 0); ptype=request.form.get("payment_type","Cash")
        if ptype=="Free": total=discount=rec=0
        bal=max(total-discount-rec,0)
        inv=Invoice(patient_id=p.id,visit_id=visit.id,branch_id=branch_id,total=total,discount=discount,received=rec,balance=bal,payment_type=ptype,created_by=u.id)
        db.session.add(inv); db.session.flush(); inv.invoice_no=f"MS-INV-{inv.id:05d}"; db.session.commit()
        log_action("ADD_PATIENT",f"{p.patient_code} / {p.name} / {inv.invoice_no}",branch_id); flash("Patient and bill saved.","success"); return redirect(url_for("patient_detail",pid=p.id))
    return render_template("patient_new.html",branches=branches,doctors=doctors,treatments=treatments)

@app.route("/patients/<int:pid>")
@login_required
def patient_detail(pid):
    p=db.session.get(Patient,pid)
    if not p or not assert_branch(p.branch_id): flash("Access denied.","danger"); return redirect(url_for("patients"))
    visits=Visit.query.filter_by(patient_id=pid).order_by(Visit.id.desc()).all(); invs=Invoice.query.filter_by(patient_id=pid).order_by(Invoice.id.desc()).all()
    return render_template("patient_detail.html",p=p,visits=visits,invs=invs)

@app.route("/patients/<int:pid>/edit",methods=["GET","POST"])
@login_required
@roles_required(ROLE_SUPER,ROLE_BRANCH)
def patient_edit(pid):
    p=db.session.get(Patient,pid)
    if not p or not assert_branch(p.branch_id): flash("Access denied.","danger"); return redirect(url_for("patients"))
    if request.method=="POST":
        p.name=request.form["name"].strip(); p.mobile=request.form.get("mobile",""); p.age=request.form.get("age",""); p.gender=request.form.get("gender",""); p.address=request.form.get("address",""); p.medical_history=request.form.get("medical_history","")
        db.session.commit(); log_action("EDIT_PATIENT",p.patient_code,p.branch_id); flash("Patient updated.","success"); return redirect(url_for("patient_detail",pid=pid))
    return render_template("patient_edit.html",p=p)

@app.route("/payments/<int:invoice_id>",methods=["POST"])
@login_required
@roles_required(ROLE_SUPER,ROLE_BRANCH,ROLE_ENTRY)
def payment_add(invoice_id):
    inv=db.session.get(Invoice,invoice_id)
    if not inv or not assert_branch(inv.branch_id): flash("Access denied.","danger"); return redirect(url_for("dashboard"))
    amt=float(request.form.get("amount") or 0); inv.received+=amt; inv.balance=max((inv.total-inv.discount)-inv.received,0)
    db.session.add(Payment(invoice_id=inv.id,amount=amt,received_by=current_user().id)); db.session.commit(); log_action("ADD_PAYMENT",f"{inv.invoice_no}: {amt}",inv.branch_id)
    return redirect(url_for("patient_detail",pid=inv.patient_id))

@app.route("/credit")
@login_required
def credit():
    bids=allowed_branch_ids(current_user()); rows=Invoice.query.filter(Invoice.branch_id.in_(bids),Invoice.balance>0).order_by(Invoice.balance.desc()).all() if bids else []
    return render_template("credit.html",rows=rows)

@app.route("/expenses",methods=["GET","POST"])
@login_required
def expenses():
    u=current_user(); bids=allowed_branch_ids(u); branches=Branch.query.filter(Branch.id.in_(bids)).all()
    if request.method=="POST":
        if u.role==ROLE_VIEWER: flash("Viewer cannot add expenses.","danger"); return redirect(url_for("expenses"))
        bid=int(request.form["branch_id"])
        if not assert_branch(bid): flash("Access denied.","danger"); return redirect(url_for("expenses"))
        e=Expense(branch_id=bid,expense_date=datetime.strptime(request.form["expense_date"],"%Y-%m-%d").date(),category=request.form.get("category",""),description=request.form.get("description",""),amount=float(request.form.get("amount") or 0),created_by=u.id)
        db.session.add(e); db.session.commit(); log_action("ADD_EXPENSE",f"{e.category}: {e.amount}",bid)
    rows=Expense.query.filter(Expense.branch_id.in_(bids)).order_by(Expense.id.desc()).limit(500).all() if bids else []
    return render_template("expenses.html",branches=branches,rows=rows,today=date.today().isoformat())

@app.route("/users",methods=["GET","POST"])
@super_required
def users():
    branches=Branch.query.filter_by(active=True).all()
    if request.method=="POST":
        email=request.form["email"].strip().lower()
        if User.query.filter_by(email=email).first(): flash("Email already exists.","danger")
        else:
            role=request.form["role"]; bid=int(request.form["branch_id"]) if request.form.get("branch_id") else None
            if role!=ROLE_SUPER and not bid: flash("Branch required.","danger"); return redirect(url_for("users"))
            u=User(name=request.form["name"].strip(),email=email,password_hash=generate_password_hash(request.form["password"]),role=role,branch_id=bid)
            db.session.add(u); db.session.commit(); log_action("ADD_USER",f"{email}/{role}",bid)
    return render_template("users.html",users=User.query.order_by(User.id.desc()).all(),branches=branches,roles=[ROLE_SUPER,ROLE_BRANCH,ROLE_ENTRY,ROLE_VIEWER])


@app.route("/users/<int:uid>/edit", methods=["GET","POST"])
@super_required
def user_edit(uid):
    me = current_user()
    target = db.session.get(User, uid)
    if not target:
        flash("User not found.", "danger")
        return redirect(url_for("users"))
    branches = Branch.query.filter_by(active=True).order_by(Branch.name).all()

    if request.method == "POST":
        email = request.form["email"].strip().lower()
        duplicate = User.query.filter(User.email == email, User.id != target.id).first()
        if duplicate:
            flash("This email is already used by another user.", "danger")
            return redirect(url_for("user_edit", uid=uid))

        target.name = request.form["name"].strip()
        target.email = email
        if request.form.get("password", "").strip():
            target.password_hash = generate_password_hash(request.form["password"].strip())

        if target.id == me.id:
            target.role = ROLE_SUPER
            target.branch_id = None
            target.active = True
        else:
            role = request.form["role"]
            bid = int(request.form["branch_id"]) if request.form.get("branch_id") else None
            if role != ROLE_SUPER and not bid:
                flash("A branch is required for this role.", "danger")
                return redirect(url_for("user_edit", uid=uid))
            target.role = role
            target.branch_id = None if role == ROLE_SUPER else bid
            target.active = request.form.get("active") == "1"

        db.session.commit()
        log_action("EDIT_USER", f"{target.email} / {target.role}", target.branch_id)
        flash("User updated successfully.", "success")
        return redirect(url_for("users"))

    return render_template("user_edit.html", target=target, branches=branches,
                           roles=[ROLE_SUPER, ROLE_BRANCH, ROLE_ENTRY, ROLE_VIEWER],
                           is_self=(target.id == me.id))

@app.route("/profile", methods=["GET","POST"])
@login_required
def profile():
    u = current_user()
    if request.method == "POST":
        email = request.form["email"].strip().lower()
        duplicate = User.query.filter(User.email == email, User.id != u.id).first()
        if duplicate:
            flash("This email is already used by another user.", "danger")
            return redirect(url_for("profile"))
        u.name = request.form["name"].strip()
        u.email = email
        if request.form.get("password", "").strip():
            u.password_hash = generate_password_hash(request.form["password"].strip())
        db.session.commit()
        log_action("EDIT_PROFILE", u.email, u.branch_id)
        flash("Your profile has been updated.", "success")
        return redirect(url_for("profile"))
    return render_template("profile.html", u=u)

@app.route("/free-patients")
@login_required
def free_patients():
    u = current_user()
    bids = allowed_branch_ids(u)
    branch_id = request.args.get("branch_id", type=int)
    q = request.args.get("q", "").strip()

    query = Invoice.query.join(Patient).filter(
        Invoice.branch_id.in_(bids),
        db.or_(Invoice.payment_type == "Free", (Invoice.total - Invoice.discount) <= 0)
    ) if bids else Invoice.query.filter(False)

    if branch_id and branch_id in bids:
        query = query.filter(Invoice.branch_id == branch_id)
    if q:
        query = query.filter(
            db.or_(
                Patient.name.ilike(f"%{q}%"),
                Patient.mobile.ilike(f"%{q}%"),
                Patient.patient_code.ilike(f"%{q}%")
            )
        )

    rows = query.order_by(Invoice.created_at.desc()).all()
    branches = Branch.query.filter(Branch.id.in_(bids)).order_by(Branch.name).all() if bids else []
    return render_template("free_patients.html", rows=rows, branches=branches,
                           selected_branch=branch_id, q=q)

@app.route("/branch-reports")
@login_required
def branch_reports():
    u = current_user()
    bids = allowed_branch_ids(u)
    branches = Branch.query.filter(Branch.id.in_(bids)).order_by(Branch.name).all() if bids else []

    branch_id = request.args.get("branch_id", type=int)
    start = request.args.get("start", "").strip()
    end = request.args.get("end", "").strip()

    target_ids = [branch_id] if branch_id and branch_id in bids else bids
    start_dt = datetime.strptime(start, "%Y-%m-%d") if start else None
    end_dt = datetime.strptime(end, "%Y-%m-%d") if end else None

    rows = []
    for b in Branch.query.filter(Branch.id.in_(target_ids)).order_by(Branch.name).all() if target_ids else []:
        pq = Patient.query.filter(Patient.branch_id == b.id)
        iq = Invoice.query.filter(Invoice.branch_id == b.id)
        eq = Expense.query.filter(Expense.branch_id == b.id)

        if start_dt:
            pq = pq.filter(Patient.created_at >= start_dt)
            iq = iq.filter(Invoice.created_at >= start_dt)
            eq = eq.filter(Expense.expense_date >= start_dt.date())
        if end_dt:
            end_exclusive = end_dt.replace(hour=23, minute=59, second=59)
            pq = pq.filter(Patient.created_at <= end_exclusive)
            iq = iq.filter(Invoice.created_at <= end_exclusive)
            eq = eq.filter(Expense.expense_date <= end_dt.date())

        invs = iq.all()
        exps = eq.all()
        free_count = iq.filter(
            db.or_(Invoice.payment_type == "Free", (Invoice.total - Invoice.discount) <= 0)
        ).count()

        billing = sum(i.total - i.discount for i in invs)
        received = sum(i.received for i in invs)
        due = sum(i.balance for i in invs)
        expenses_total = sum(e.amount for e in exps)

        rows.append({
            "id": b.id,
            "name": b.name,
            "patients": pq.count(),
            "billing": billing,
            "received": received,
            "due": due,
            "expenses": expenses_total,
            "net": received - expenses_total,
            "free": free_count,
        })

    return render_template("branch_reports.html", rows=rows, branches=branches,
                           selected_branch=branch_id, start=start, end=end)

@app.route("/import-excel", methods=["GET","POST"])
@super_required
def import_excel():
    if request.method == "POST":
        uploaded = request.files.get("excel_file")
        if not uploaded or not uploaded.filename:
            flash("Please choose an Excel file.", "danger")
            return redirect(url_for("import_excel"))
        if not uploaded.filename.lower().endswith((".xlsx", ".xlsm")):
            flash("Only .xlsx or .xlsm files are supported.", "danger")
            return redirect(url_for("import_excel"))

        tmp_name = f"import_{int(datetime.utcnow().timestamp())}_{secure_filename(uploaded.filename)}"
        tmp_path = os.path.join(app.config["UPLOAD_FOLDER"], tmp_name)
        uploaded.save(tmp_path)

        added_patients = 0
        skipped_patients = 0
        added_expenses = 0
        try:
            wb = load_workbook(tmp_path, data_only=True, read_only=True)

            if "Patient Database" not in wb.sheetnames:
                raise ValueError("Patient Database sheet was not found.")

            ws = wb["Patient Database"]
            for r in range(6, ws.max_row + 1):
                serial = ws.cell(r, 1).value
                patient_name = ws.cell(r, 4).value
                branch_name = ws.cell(r, 3).value
                if not patient_name or not branch_name:
                    continue

                serial_text = str(int(serial)) if isinstance(serial, (int, float)) else str(serial or r-5).strip()
                import_code = f"MS-XLS-{serial_text.zfill(5)}"
                if Patient.query.filter_by(patient_code=import_code).first():
                    skipped_patients += 1
                    continue

                b = get_or_create_branch(str(branch_name))
                total = float(ws.cell(r, 23).value or 0)
                received = float(ws.cell(r, 24).value or 0)
                due_raw = ws.cell(r, 25).value
                due = float(due_raw if due_raw not in (None, "") else max(total - received, 0))
                payment = str(ws.cell(r, 26).value or ("Free" if total <= 0 else "Cash"))
                doctor_name = str(ws.cell(r, 27).value or "").strip()

                doctor = None
                if doctor_name:
                    doctor = Doctor.query.filter(
                        db.func.lower(Doctor.name) == doctor_name.lower(),
                        Doctor.branch_id == b.id
                    ).first()
                    if not doctor:
                        doctor = Doctor(name=doctor_name, branch_id=b.id)
                        db.session.add(doctor)
                        db.session.flush()

                p = Patient(
                    patient_code=import_code,
                    branch_id=b.id,
                    name=str(patient_name).strip(),
                    mobile=str(ws.cell(r, 7).value or ""),
                    age=str(ws.cell(r, 5).value or ""),
                    gender=str(ws.cell(r, 6).value or ""),
                    address=str(ws.cell(r, 29).value or ""),
                    medical_history="Imported from Excel",
                    created_by=current_user().id,
                    created_at=datetime.combine(excel_date(ws.cell(r, 2).value), datetime.min.time())
                )
                db.session.add(p)
                db.session.flush()

                treatment_notes = []
                first_treatment = None
                for col, tname in EXCEL_TREATMENT_COLUMNS.items():
                    val = ws.cell(r, col).value
                    if val not in (None, ""):
                        t = get_or_create_treatment(tname)
                        if first_treatment is None:
                            first_treatment = t
                        treatment_notes.append(f"{tname}: {val}")

                visit = Visit(
                    patient_id=p.id,
                    branch_id=b.id,
                    doctor_id=doctor.id if doctor else None,
                    treatment_id=first_treatment.id if first_treatment else None,
                    tooth_notes=" | ".join(treatment_notes),
                    clinical_notes="Imported from Excel Patient Database",
                    visit_date=excel_date(ws.cell(r, 2).value),
                    created_by=current_user().id
                )
                db.session.add(visit)
                db.session.flush()

                inv = Invoice(
                    patient_id=p.id,
                    visit_id=visit.id,
                    branch_id=b.id,
                    total=total,
                    discount=0,
                    received=received,
                    balance=max(due, 0),
                    payment_type="Free" if total <= 0 else payment,
                    created_by=current_user().id,
                    created_at=datetime.combine(excel_date(ws.cell(r, 2).value), datetime.min.time())
                )
                db.session.add(inv)
                db.session.flush()
                inv.invoice_no = f"MS-XLS-INV-{p.id:05d}"
                added_patients += 1

            if "Daily Expenses" in wb.sheetnames:
                ws = wb["Daily Expenses"]
                for r in range(8, ws.max_row + 1):
                    exp_date = ws.cell(r, 2).value
                    branch_name = ws.cell(r, 3).value
                    amount = ws.cell(r, 7).value
                    if not exp_date or not branch_name or amount in (None, ""):
                        continue

                    b = get_or_create_branch(str(branch_name))
                    # Simple duplicate guard by exact main fields.
                    d = excel_date(exp_date)
                    category = str(ws.cell(r, 4).value or "")
                    description = str(ws.cell(r, 5).value or "")
                    amount_f = float(amount or 0)
                    exists = Expense.query.filter_by(
                        branch_id=b.id, expense_date=d, category=category,
                        description=description, amount=amount_f
                    ).first()
                    if not exists:
                        db.session.add(Expense(
                            branch_id=b.id,
                            expense_date=d,
                            category=category,
                            description=description,
                            amount=amount_f,
                            created_by=current_user().id
                        ))
                        added_expenses += 1

            db.session.commit()
            log_action("IMPORT_EXCEL",
                       f"Patients added: {added_patients}; skipped: {skipped_patients}; expenses added: {added_expenses}",
                       None)
            flash(f"Import complete: {added_patients} patients added, {skipped_patients} skipped, {added_expenses} expenses added.", "success")
        except Exception as e:
            db.session.rollback()
            flash(f"Excel import failed: {e}", "danger")
        finally:
            try:
                os.remove(tmp_path)
            except OSError:
                pass
        return redirect(url_for("import_excel"))

    return render_template("import_excel.html")


@app.route("/branches",methods=["GET","POST"])
@super_required
def branches():
    if request.method=="POST":
        b=Branch(name=request.form["name"].strip(),country=request.form.get("country","Pakistan"),city=request.form.get("city",""),address=request.form.get("address",""),phone=request.form.get("phone",""))
        db.session.add(b); db.session.commit(); log_action("ADD_BRANCH",b.name,b.id)
    return render_template("branches.html",rows=Branch.query.order_by(Branch.id).all())

@app.route("/doctors",methods=["GET","POST"])
@login_required
def doctors():
    u=current_user(); bids=allowed_branch_ids(u); branches=Branch.query.filter(Branch.id.in_(bids)).all()
    if request.method=="POST":
        if u.role not in (ROLE_SUPER,ROLE_BRANCH): flash("Permission denied.","danger"); return redirect(url_for("doctors"))
        bid=int(request.form["branch_id"])
        if not assert_branch(bid): flash("Access denied.","danger"); return redirect(url_for("doctors"))
        db.session.add(Doctor(name=request.form["name"].strip(),mobile=request.form.get("mobile",""),branch_id=bid)); db.session.commit()
    rows=Doctor.query.filter(Doctor.branch_id.in_(bids)).all() if bids else []
    return render_template("doctors.html",rows=rows,branches=branches)

@app.route("/staff",methods=["GET","POST"])
@login_required
def staff():
    u=current_user(); bids=allowed_branch_ids(u); branches=Branch.query.filter(Branch.id.in_(bids)).all()
    if request.method=="POST":
        if u.role not in (ROLE_SUPER,ROLE_BRANCH): flash("Permission denied.","danger"); return redirect(url_for("staff"))
        bid=int(request.form["branch_id"])
        if not assert_branch(bid): flash("Access denied.","danger"); return redirect(url_for("staff"))
        s=Staff(name=request.form["name"].strip(),role_name=request.form.get("role_name",""),mobile=request.form.get("mobile",""),cnic=request.form.get("cnic",""),branch_id=bid,salary=float(request.form.get("salary") or 0))
        db.session.add(s); db.session.commit()
    rows=Staff.query.filter(Staff.branch_id.in_(bids)).all() if bids else []
    return render_template("staff.html",rows=rows,branches=branches)

@app.route("/salaries",methods=["GET","POST"])
@login_required
def salaries():
    u=current_user(); bids=allowed_branch_ids(u); staffs=Staff.query.filter(Staff.branch_id.in_(bids)).all() if bids else []
    if request.method=="POST":
        if u.role not in (ROLE_SUPER,ROLE_BRANCH): flash("Permission denied.","danger"); return redirect(url_for("salaries"))
        sid=int(request.form["staff_id"]); s=db.session.get(Staff,sid)
        if not s or not assert_branch(s.branch_id): flash("Access denied.","danger"); return redirect(url_for("salaries"))
        gross=float(request.form.get("gross_salary") or s.salary or 0); adv=float(request.form.get("advance") or 0); paid=float(request.form.get("paid") or 0)
        db.session.add(SalaryPayment(staff_id=sid,month=request.form["month"],gross_salary=gross,advance=adv,paid=paid,balance=max(gross-adv-paid,0),paid_date=date.today())); db.session.commit()
    rows=SalaryPayment.query.join(Staff).filter(Staff.branch_id.in_(bids)).order_by(SalaryPayment.id.desc()).all() if bids else []
    return render_template("salaries.html",rows=rows,staffs=staffs)

@app.route("/treatments",methods=["GET","POST"])
@super_required
def treatments():
    if request.method=="POST":
        name=request.form["name"].strip(); price=float(request.form.get("price") or 0); t=Treatment.query.filter_by(name=name).first()
        if t: t.price=price
        else: db.session.add(Treatment(name=name,price=price))
        db.session.commit()
    return render_template("treatments.html",rows=Treatment.query.order_by(Treatment.name).all())

@app.route("/activity")
@super_required
def activity():
    return render_template("activity.html",rows=ActivityLog.query.order_by(ActivityLog.id.desc()).limit(1000).all())

@app.route("/reports")
@login_required
def reports():
    u=current_user(); bids=allowed_branch_ids(u)
    invs=Invoice.query.filter(Invoice.branch_id.in_(bids)).all() if bids else []
    exps=Expense.query.filter(Expense.branch_id.in_(bids)).all() if bids else []
    treatments={}
    doctors={}
    months={}
    for inv in invs:
        t=inv.visit.treatment.name if inv.visit and inv.visit.treatment else "Other"; treatments[t]=treatments.get(t,0)+(inv.total-inv.discount)
        d=inv.visit.doctor.name if inv.visit and inv.visit.doctor else "Unassigned"; doctors[d]=doctors.get(d,0)+(inv.total-inv.discount)
        m=inv.created_at.strftime("%Y-%m"); months[m]=months.get(m,0)+inv.received
    return render_template("reports.html",treatments=treatments,doctors=doctors,months=months,
                           billing=sum(i.total-i.discount for i in invs),received=sum(i.received for i in invs),due=sum(i.balance for i in invs),expenses=sum(e.amount for e in exps))

@app.route("/export/patients.csv")
@login_required
def export_patients():
    bids=allowed_branch_ids(current_user()); rows=Patient.query.filter(Patient.branch_id.in_(bids)).all() if bids else []
    s=BytesIO(); text="Patient ID,Name,Mobile,Branch,Created\n"+"\n".join([f"{p.patient_code},{p.name},{p.mobile},{p.branch.name},{p.created_at.date()}" for p in rows]); s.write(text.encode("utf-8-sig")); s.seek(0)
    return send_file(s,mimetype="text/csv",as_attachment=True,download_name="patients.csv")

@app.route("/backup")
@super_required
def backup():
    if not app.config["SQLALCHEMY_DATABASE_URI"].startswith("sqlite"):
        flash("For PostgreSQL use your host's database backup/snapshot.","success"); return redirect(url_for("dashboard"))
    dbfile=os.path.join(os.path.dirname(__file__),"instance","ms_dental_final.db")
    if not os.path.exists(dbfile): flash("Database file not found.","danger"); return redirect(url_for("dashboard"))
    return send_file(dbfile,as_attachment=True,download_name=f"MS_Dental_Backup_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.db")

def seed():
    db.create_all()
    if Branch.query.count()==0:
        db.session.add_all([Branch(name="Main Branch",country="Pakistan"),Branch(name="Second Branch",country="Pakistan"),Branch(name="Third Branch",country="Pakistan")]); db.session.commit()
    if User.query.count()==0:
        db.session.add(User(name="Super Admin",email="admin@msdental.local",password_hash=generate_password_hash("1234"),role=ROLE_SUPER))
    if Treatment.query.count()==0:
        for n in ["Consultation","Scaling & Polishing","Composite Filling","GIC Filling","Tooth Extraction","Surgical Extraction","Root Canal Treatment","Crown","Bridge","Denture","Implant","Braces","X-Ray","Other"]:
            db.session.add(Treatment(name=n,price=0))
    db.session.commit()

with app.app_context(): seed()

if __name__=="__main__":
    app.run(host="0.0.0.0",port=int(os.environ.get("PORT",5000)),debug=os.environ.get("FLASK_DEBUG")=="1")
